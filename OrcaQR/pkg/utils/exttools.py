import os
import re
import xlrd
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string


def get_worksheets(file: str, use_xls: bool) -> str:
    if not use_xls:
        xl = load_workbook(file, data_only=True)
        return xl.worksheets, xl.sheetnames
    else:
        xl = xlrd.open_workbook(file)
        return xl.sheets(), xl.sheet_names()


def get_headers(sheet: object, use_xls: bool) -> dict:
    header_row = sheet[1] if not use_xls else sheet.row(0)
    first_index = 1 if not use_xls else 0
    header_dict = {
        header.value: column_index
        for column_index, header in enumerate(header_row, first_index)
        if header.value is not None
    }
    return header_dict


def read_excel(
    use_xls: bool,
    sheet: object,
    header_dict: dict,
    regex: str,
    label_data: list,
    show_cols: dict,
) -> list[list[str]]:
    skip_cols = []
    assert label_data is not None, "No label data specified"
    column_indices = [
        header_dict.get(column_name)
        for column_name, show_bool in show_cols.items()
        if show_bool
    ]
    column_indices.extend([header_dict.get(data) for data in label_data])
    column_indices.sort()

    if not column_indices:
        raise AssertionError(
            "No columns are matched to show, check Settings: Columns to show,\n Caution, values are case sensitive"
        )
    all_row_vals = []
    if not use_xls:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            sel_col_val = {}
            for col_index in column_indices:
                cell_value = row[col_index - 1]
                if isinstance(cell_value, float):
                    cell_value = int(cell_value)
                if not re.match(re.compile(regex), str(cell_value)):
                    skip_col = list(header_dict.keys())[
                        list(header_dict.values()).index(col_index)
                    ]
                    if skip_col not in skip_cols:
                        skip_cols.append(str(skip_col))
                    show_cols.update({skip_col: False})
                    continue
                sel_col_val.update(
                    {
                        list(header_dict.keys())[
                            list(header_dict.values()).index(col_index)
                        ]: str(cell_value)
                    }
                )

            all_row_vals.append(sel_col_val)
    else:
        for row_index in range(1, sheet.nrows):  # Start from row 2
            row = sheet.row(row_index)
            sel_col_val = []
            for col_index in column_indices:
                cell_value = row[col_index].value
                if isinstance(cell_value, float):
                    cell_value = int(cell_value)
                if not re.match(re.compile(regex), str(cell_value)):
                    skip_col = list(header_dict.keys())[
                        list(header_dict.values()).index(col_index)
                    ]
                    if skip_col not in skip_cols:
                        skip_cols.append(str(skip_col))
                    show_cols.update({skip_col: False})
                    continue
                sel_col_val.update(
                    {
                        list(header_dict.keys())[
                            list(header_dict.values()).index(col_index)
                        ]: str(cell_value)
                    }
                )
            all_row_vals.append(sel_col_val)

    skip_set = set(skip_cols)
    _ic = [x for x in label_data if x not in skip_set]
    data_not_incl = [x for x in label_data if x in skip_set]
    data_not_incl = set(data_not_incl)
    label_data = _ic[:]  # remove all invalid columns from label data
    return all_row_vals, label_data, data_not_incl, show_cols


def write_excel(file, row, value):
    xl = load_workbook(file)
    sheet = xl.worksheets[0]


def get_icon_path(icon_name: str) -> str:
    base_path = os.path.abspath(".")
    return os.path.join(base_path, "resources", "icons", f"{icon_name}.jpg")


def get_font_path(font_name: str) -> str:
    base_path = os.path.abspath(".")
    return os.path.join(base_path, "resources", "fonts", f"{font_name}.ttf")


def get_label_path(label_name: str) -> str:
    base_path = os.path.abspath(".")
    return os.path.join(base_path, "resources", "fonts", f"{label_name}.png")
